#!/usr/bin/env python3
"""
build.py — Reads the Excel menu file and updates index.html automatically.

Workflow:
  1. Edit prices / items in the Excel file
  2. Run:  python build.py
  3. Run:  git add index.html && git commit -m "update prices" && git push

Change EXCEL_FILE below if you rename the spreadsheet.
New categories found in Excel are auto-detected and saved to CATEGORY_MAP here.
"""

import re
import sys
import unicodedata
from pathlib import Path
import openpyxl

# ── CONFIG ── change these if you rename files ───────────────────────────────
EXCEL_FILE = "Cafe_Menu_Data_updated.xlsx"
HTML_FILE  = "index.html"
# ─────────────────────────────────────────────────────────────────────────────

# Maps Excel category name → (section_id, bg_label, en_label, note_bg, note_en, order)
CATEGORY_MAP = {
    'Топли Напитки / Hot Drinks': ('hot', 'Топли Напитки', 'Hot Drinks', None, None,  1),
    'Топли Напитки (Ядково Мляко)': ('hot-plant', 'Топли Напитки с Ядково Мляко', 'Hot Drinks — Plant Milk', 'овес · бадем · кокос', 'oat · almond · coconut',  2),
    'Студени Напитки / Cold Drinks': ('cold', 'Студени Напитки', 'Cold Drinks', None, None,  3),
    'Айс Напитки / Iced Drinks': ('iced', 'Айс Напитки', 'Iced Drinks', None, None,  4),
    'Безалкохолни Напитки / Soft Drinks': ('soft', 'Безалкохолни Напитки', 'Soft Drinks', None, None,  5),
    'Безалкохолни Коктейли / Mocktails': ('mocktails', 'Безалкохолни Коктейли', 'Mocktails', None, None,  6),
    'Алкохолни Коктейли / Cocktails': ('cocktails', 'Алкохолни Коктейли', 'Cocktails', None, None,  7),
    'Бира / Beer': ('beer-bg', 'Бира', 'Beer', None, None,  8),
    'Бира Внос / Import Beer': ('beer-import', 'Бира Внос', 'Import Beer', None, None,  9),
    'Вино / Wine': ('wine', 'Вино', 'Wine', None, None, 10),
    'Алкохол / Alcohol': ('spirits', 'Алкохол', 'Spirits', None, None, 11),
    'Уиски / Whiskey': ('whiskey-bg', 'Уиски', 'Whiskey', None, None, 12),
    'Уйски Внос / Import Whiskey': ('whiskey-import', 'Уиски Внос', 'Import Whiskey', None, None, 13),
    'Водка & Джин / Vodka & Gin': ('vodka-gin', 'Водка & Джин', 'Vodka & Gin', None, None, 14),
    'Ром / Rum': ('rum', 'Ром', 'Rum', None, None, 15),
    'Вермут / Vermouth': ('vermouth', 'Вермут', 'Vermouth', None, None, 16),
    'Ликьори / Liqueurs': ('liqueurs', 'Ликьори', 'Liqueurs', None, None, 17),
    'Десерти / Desserts': ('desserts', 'Десерти', 'Desserts', None, None, 18),
    'Ядки / Nuts': ('nuts', 'Ядки', 'Nuts', None, None, 19),
    'Анасонови аперитиви / Anise-based Spirits': ('anise-based-spirits', 'Анасонови аперитиви', 'Anise-based Spirits', None, None, 20),
    'Коняк / Brandy': ('brandy', 'Коняк', 'Brandy', None, None, 21),
    'Ракия / Rakia': ('rakia', 'Ракия', 'Rakia', None, None, 22),
    'Текила / Tequila': ('tequila', 'Текила', 'Tequila', None, None, 23),
    'Уиски Българско / Whiskey Bulgarian': ('whiskey-bulgarian', 'Уиски Българско', 'Whiskey Bulgarian', None, None, 24),
}


# ── Auto-category helpers ─────────────────────────────────────────────────────

def _slugify(text: str) -> str:
    """Convert an English label to a URL-safe slug for use as section_id."""
    text = unicodedata.normalize("NFKD", text).encode("ascii", "ignore").decode()
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "-", text)
    return text.strip("-")


def auto_register_category(cat_raw: str) -> tuple:
    """
    Derive a CATEGORY_MAP entry from an unknown Excel category string and
    register it into CATEGORY_MAP so it is used in this run.
    The new entry is also saved back into build.py itself.
    """
    parts = cat_raw.split(" / ", 1)
    bg_label = parts[0].strip()
    en_label = parts[1].strip() if len(parts) > 1 else parts[0].strip()

    # Build a unique section_id
    base_id = _slugify(en_label) or _slugify(bg_label) or "section"
    section_id = base_id
    existing_ids = {v[0] for v in CATEGORY_MAP.values()}
    counter = 2
    while section_id in existing_ids:
        section_id = f"{base_id}-{counter}"
        counter += 1

    order = max((v[5] for v in CATEGORY_MAP.values()), default=0) + 1
    entry = (section_id, bg_label, en_label, None, None, order)
    CATEGORY_MAP[cat_raw] = entry
    return entry


def _format_category_map() -> str:
    """Regenerate the full CATEGORY_MAP source block (to be written back to build.py)."""
    lines = ["CATEGORY_MAP = {
    'Топли Напитки / Hot Drinks': ('hot', 'Топли Напитки', 'Hot Drinks', None, None,  1),
    'Топли Напитки (Ядково Мляко)': ('hot-plant', 'Топли Напитки с Ядково Мляко', 'Hot Drinks — Plant Milk', 'овес · бадем · кокос', 'oat · almond · coconut',  2),
    'Студени Напитки / Cold Drinks': ('cold', 'Студени Напитки', 'Cold Drinks', None, None,  3),
    'Айс Напитки / Iced Drinks': ('iced', 'Айс Напитки', 'Iced Drinks', None, None,  4),
    'Безалкохолни Напитки / Soft Drinks': ('soft', 'Безалкохолни Напитки', 'Soft Drinks', None, None,  5),
    'Безалкохолни Коктейли / Mocktails': ('mocktails', 'Безалкохолни Коктейли', 'Mocktails', None, None,  6),
    'Алкохолни Коктейли / Cocktails': ('cocktails', 'Алкохолни Коктейли', 'Cocktails', None, None,  7),
    'Бира / Beer': ('beer-bg', 'Бира', 'Beer', None, None,  8),
    'Бира Внос / Import Beer': ('beer-import', 'Бира Внос', 'Import Beer', None, None,  9),
    'Вино / Wine': ('wine', 'Вино', 'Wine', None, None, 10),
    'Алкохол / Alcohol': ('spirits', 'Алкохол', 'Spirits', None, None, 11),
    'Уиски / Whiskey': ('whiskey-bg', 'Уиски', 'Whiskey', None, None, 12),
    'Уйски Внос / Import Whiskey': ('whiskey-import', 'Уиски Внос', 'Import Whiskey', None, None, 13),
    'Водка & Джин / Vodka & Gin': ('vodka-gin', 'Водка & Джин', 'Vodka & Gin', None, None, 14),
    'Ром / Rum': ('rum', 'Ром', 'Rum', None, None, 15),
    'Вермут / Vermouth': ('vermouth', 'Вермут', 'Vermouth', None, None, 16),
    'Ликьори / Liqueurs': ('liqueurs', 'Ликьори', 'Liqueurs', None, None, 17),
    'Десерти / Desserts': ('desserts', 'Десерти', 'Desserts', None, None, 18),
    'Ядки / Nuts': ('nuts', 'Ядки', 'Nuts', None, None, 19),
    'Анасонови аперитиви / Anise-based Spirits': ('anise-based-spirits', 'Анасонови аперитиви', 'Anise-based Spirits', None, None, 20),
    'Коняк / Brandy': ('brandy', 'Коняк', 'Brandy', None, None, 21),
    'Ракия / Rakia': ('rakia', 'Ракия', 'Rakia', None, None, 22),
    'Текила / Tequila': ('tequila', 'Текила', 'Tequila', None, None, 23),
    'Уиски Българско / Whiskey Bulgarian': ('whiskey-bulgarian', 'Уиски Българско', 'Whiskey Bulgarian', None, None, 24),
}: "
            f"({repr(sid)}, {repr(bg)}, {repr(en)}, "
            f"{repr(note_bg)}, {repr(note_en)}, {order:>2}),"
        )
    lines.append("}")
    return "\n".join(lines)


def update_category_map_in_script() -> bool:
    """Rewrite the CATEGORY_MAP block inside this script file with any new entries."""
    script = Path(__file__)
    source = script.read_text(encoding="utf-8")
    new_block = _format_category_map()
    new_source = re.sub(
        r"CATEGORY_MAP = \{.*?\}",
        new_block,
        source,
        flags=re.DOTALL,
    )
    if new_source == source:
        print("  ⚠  Could not save new categories back to build.py", file=sys.stderr)
        return False
    script.write_text(new_source, encoding="utf-8")
    return True


# ── Menu builder ─────────────────────────────────────────────────────────────

def js_str(s):
    """Return a JS single-quoted string literal, properly escaped."""
    s = s.replace("\\", "\\\\").replace("'", "\\'")
    return f"'{s}'"


def build_menu_js(excel_path: Path) -> str:
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active

    # Columns: Category | Description/Ingredients | Name(EN) | Name(BG) | Volume | BGN-formula | EUR
    #          0          1                          2           3          4        5              6

    from collections import OrderedDict
    cats: dict[str, list] = OrderedDict()

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        if not any(v is not None for v in row):
            continue
        cat_raw, desc, name_en, name_bg, vol, _bgn, eur = row
        if not cat_raw:
            continue
        cat_raw = str(cat_raw).strip()

        if cat_raw not in CATEGORY_MAP:
            auto_register_category(cat_raw)
            print(f"  + New category detected and added: '{cat_raw}'")

        if cat_raw not in cats:
            cats[cat_raw] = []

        # Split bilingual description  "bg part / en part"
        dbg = den = None
        if desc:
            parts = str(desc).split(" / ", 1)
            dbg = parts[0].strip() or None
            den = parts[1].strip() if len(parts) > 1 else None

        cats[cat_raw].append({
            "bg":  str(name_bg).strip() if name_bg else "",
            "en":  str(name_en).strip() if name_en else "",
            "vol": str(vol).strip() if vol else None,
            "eur": float(eur) if eur is not None else None,
            "dbg": dbg,
            "den": den,
        })

    # Sort by the display order defined in CATEGORY_MAP
    sorted_cats = sorted(cats.items(), key=lambda kv: CATEGORY_MAP[kv[0]][5])

    lines = ["const MENU=["]
    for cat_raw, items in sorted_cats:
        sid, bg_lbl, en_lbl, note_bg, note_en, _ = CATEGORY_MAP[cat_raw]
        line = f"{{id:{js_str(sid)},bg:{js_str(bg_lbl)},en:{js_str(en_lbl)},"
        if note_bg:
            line += f"note:{{bg:{js_str(note_bg)},en:{js_str(note_en)}}},"
        line += "items:["

        item_parts = []
        for item in items:
            p = f"{{bg:{js_str(item['bg'])},en:{js_str(item['en'])}"
            if item["dbg"]: p += f",dbg:{js_str(item['dbg'])}"
            if item["den"]: p += f",den:{js_str(item['den'])}"
            if item["vol"]: p += f",vol:{js_str(item['vol'])}"
            if item["eur"] is not None:
                p += f",eur:{item['eur']:.2f}"
            else:
                p += ",eur:null"
            p += "}"
            item_parts.append(p)

        line += ",".join(item_parts) + "]},"
        lines.append(line)

    lines.append("];")
    return "\n".join(lines)


def inject_into_html(html_path: Path, menu_js: str):
    html = html_path.read_text(encoding="utf-8")
    pattern = r"const MENU=\[.*?\];"
    if not re.search(pattern, html, flags=re.DOTALL):
        print("ERROR: Could not find 'const MENU=[...]' block in index.html", file=sys.stderr)
        sys.exit(1)
    new_html = re.sub(pattern, menu_js, html, flags=re.DOTALL)
    html_path.write_text(new_html, encoding="utf-8")


if __name__ == "__main__":
    base = Path(__file__).parent
    excel_path = base / EXCEL_FILE
    html_path  = base / HTML_FILE

    if not excel_path.exists():
        print(f"ERROR: Excel file not found: {excel_path}\n"
              f"Set EXCEL_FILE at the top of build.py to match your filename.", file=sys.stderr)
        sys.exit(1)
    if not html_path.exists():
        print(f"ERROR: {HTML_FILE} not found in {base}", file=sys.stderr)
        sys.exit(1)

    known_before = set(CATEGORY_MAP.keys())

    print(f"Reading  {excel_path.name} ...")
    menu_js = build_menu_js(excel_path)
    item_count = menu_js.count("eur:")
    print(f"  → {item_count} items across {menu_js.count('id:')} categories")

    new_cats = [k for k in CATEGORY_MAP if k not in known_before]
    if new_cats:
        n = len(new_cats)
        print(f"  → Saving {n} new categor{'y' if n == 1 else 'ies'} to build.py ...")
        update_category_map_in_script()

    print(f"Updating {html_path.name} ...")
    inject_into_html(html_path, menu_js)

    print("\nDone! To publish:")
    print("  git add index.html")
    print('  git commit -m "update prices"')
    print("  git push")
